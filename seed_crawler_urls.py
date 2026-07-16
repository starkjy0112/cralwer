# -*- coding: utf-8 -*-
"""xlsx 13-43행 (검증 완료) 크롤러 URL을 DB에 seed.
matplotlib 의존성 없이 crawler_registry.py에서 로드.
"""
import os
import db
from crawler_registry import VERIFIED_ROWS

# URL 정보는 xlsx가 있으면 거기서, 없으면 하드코딩된 fallback URL 사용
FALLBACK_URLS = {
    # 13-43 검증 완료 크롤러의 URL (xlsx 파일 없어도 seed 가능)
    "alio_item":   ("공고", "https://alio.go.kr/item/itemOrganList.do?apbaId=C0107"),
    "nara":        ("입찰공고", "https://www.g2b.go.kr/"),
    "alio":        ("공고", "https://www.alio.go.kr/occasional/bidList.do?"),
    "lh":          ("자재공법심의", "https://partner.lh.or.kr/deliberate/deliberate.asp"),
    "kr":          ("공지사항", "https://www.kr.or.kr/boardCnts/list.do?boardID=51"),
    "ekr":         ("공지사항", "https://www.ekr.or.kr/planweb/board/list.krc"),
    "gtdc":        ("입찰공고", "https://gtdc.or.kr/pub/egbid.do"),
    "gdco":        ("입찰공고", "https://www.gdco.co.kr/customer/bidding_list.php?strBoardID=BID"),
    "gmdc":        ("입찰공고", "https://www.gmdc.co.kr/bbs/board.php?bo_table=sub04_02"),
    "gndc":        ("공고", "https://www.gndc.co.kr/boardlist.do?seqId=0000000048"),
    "gbdc":        ("게시판 검색", "https://www.gbdc.co.kr/totalSearch.do?searchKeywordTotal=&pageIndex=1&tapIdx=2"),
    "ghdc":        ("통합검색", "https://ghdc.or.kr/sub.html?code=08_05"),
    "dudc":        ("공지사항", "https://www.dudc.or.kr/ko/page.do?mnu_uid=100"),
    "sdco":        ("고시/공고", "https://www.sdco.or.kr/board.es?mid=a10601020000&bid=0007"),
    "sh":          ("통합검색", "https://www.i-sh.co.kr/search/total"),
    "sh_bid":      ("입찰공고", "https://www.i-sh.co.kr/main/lay2/program/S1T316C7212/www/m_2428/BidblancList.do"),
    "isdc":        ("통합검색", "https://www.isdc.co.kr/guidance/search.asp"),
    "isdc_notice": ("고시공고", "https://www.isdc.co.kr/board/default/boardDefaultList.asp?HiddenBbsNo=82"),
    "jndc":        ("통합검색", "https://www.jndc.co.kr/cf/search.do"),
    "jbdc":        ("통합검색", "https://www.jbdc.co.kr/search/board.do?searchWrd="),
    "jpdc":        ("검색서비스", "https://www.jpdc.co.kr/help/search.htm?q="),
    "cbdc":        ("공지사항", "https://www.cbdc.co.kr/zboard/list.do?lmCode=BBSMSTR_000000000018"),
    "cndc":        ("입찰공고", "https://www.cndc.kr/bbs/list.do?key=2404080009"),
    "ttdc":        ("입찰정보", "http://corp.ttdc.kr/board/board.aspx?tbl=bidding"),
    "gcuc":        ("개발사업", "https://www.gcuc.or.kr/fmcs/722"),
    "gmcc":        ("통합검색", "https://www.gmcc.co.kr/findeepSearch.es?mid=a10603000000"),
    "gh":          ("기술제안공고", "https://www.gh.or.kr/gh/technical-suggestion-notice-list.do"),
    "gys":         ("입찰공고/고시", "https://www.gys.or.kr/subpage/index/46"),
    "guriuc":      ("입찰정보", "https://www.guriuc.or.kr/bbsArticle/list.do?bbsId=BID_INFO"),
    "gunpouc":     ("입찰내역", "https://www.gunpouc.or.kr/fmcs/90"),
    "ncuc":        ("공유재산입찰", "https://www.ncuc.or.kr/main/35"),
}

# 폴백 URL (기존 URL 죽었을 때 대체)
FALLBACK_URL_MAP = {
    "jndc": "https://www.jndc.co.kr/web/main/searchResult",
    "gndc": "https://www.gndc.co.kr/boardlist.do?seqId=0000006241",
}


def seed():
    db.init_db()

    # xlsx 파일 있으면 거기서 우선 로드, 없으면 FALLBACK_URLS 사용
    xlsx_data = {}
    xlsx_path = "서칭기관 목록(250311)_중복삭제.xlsx"
    if os.path.exists(xlsx_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            for r in range(13, 57):
                cid_row = None
                for row_num, cid, _, _ in VERIFIED_ROWS:
                    if row_num == r:
                        cid_row = cid
                        break
                if cid_row:
                    path = (ws.cell(r, 4).value or "").strip()
                    url = (ws.cell(r, 5).value or "").strip()
                    board_name = path.split(">")[-1].strip() if ">" in path else (path or "게시판")
                    if url:
                        xlsx_data[cid_row] = (board_name, url)
        except Exception as e:
            print(f"  ⚠ xlsx 로드 실패, FALLBACK_URLS 사용: {e}")

    added = 0
    for row_num, cid, name, policy in VERIFIED_ROWS:
        # xlsx 우선, 없으면 FALLBACK_URLS
        board_name, url = xlsx_data.get(cid) or FALLBACK_URLS.get(cid, (None, None))
        if not url:
            print(f"  ⚠ 행 {row_num} {cid}: URL 없음, skip")
            continue

        db.upsert_crawler_url(
            crawler_id=cid,
            crawler_name=name,
            board_key="default",
            board_name=board_name or "게시판",
            url=url,
            fallback_url=FALLBACK_URL_MAP.get(cid),
            is_active=1,
            priority=10,
            notes=f"xlsx {row_num}행 검증 완료",
        )
        added += 1

    print(f"\n등록: {added}개")

    urls = db.get_crawler_urls()
    print(f"DB 총 URL: {len(urls)}")

    verified = db.get_verified_crawler_ids()
    print(f"검증된 크롤러: {len(verified)}개")


if __name__ == "__main__":
    seed()
